"""
Excel 生成辅助工具

从 drama-generator-pro 技能的结构化数据生成三张标准化 Excel 表格：
- 人物信息表
- 场景信息表
- 分镜信息表

基于 skill技能/drama-generator-pro-1.0.0/scripts/ 下的脚本逻辑重写为可直接调用的模块。
依赖：openpyxl
"""
import logging
import os
import time
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    logger.warning("[ExcelGenerator] openpyxl 未安装，Excel 生成功能将不可用")


# ─── 样式常量 ───────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if _OPENPYXL_AVAILABLE else None
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11) if _OPENPYXL_AVAILABLE else None
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
) if _OPENPYXL_AVAILABLE else None
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top") if _OPENPYXL_AVAILABLE else None
_CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center") if _OPENPYXL_AVAILABLE else None
_CENTER_TOP_ALIGNMENT = Alignment(horizontal="center", vertical="top") if _OPENPYXL_AVAILABLE else None


def _apply_header_style(cell) -> None:
    """应用表头样式"""
    if not _OPENPYXL_AVAILABLE:
        return
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.border = _THIN_BORDER
    cell.alignment = _CENTER_ALIGNMENT


def _apply_cell_style(cell, wrap: bool = False, center: bool = False) -> None:
    """应用单元格样式"""
    if not _OPENPYXL_AVAILABLE:
        return
    cell.border = _THIN_BORDER
    if wrap:
        cell.alignment = _WRAP_ALIGNMENT
    elif center:
        cell.alignment = _CENTER_TOP_ALIGNMENT


# ─── 人物信息表 ─────────────────────────────────────────

def create_character_sheet(ws, characters: List[Dict[str, Any]]) -> None:
    """创建人物信息表工作表"""
    ws.title = "人物信息表"

    headers = ["人物编号", "人物姓名", "人物信息提取", "人物小传", "人物三视图提示词"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    for row, char in enumerate(characters, 2):
        ws.cell(row=row, column=1, value=char.get("id", f"C{row-1:02d}"))
        _apply_cell_style(ws.cell(row=row, column=1), center=True)

        ws.cell(row=row, column=2, value=char.get("name", ""))
        _apply_cell_style(ws.cell(row=row, column=2), center=True)

        ws.cell(row=row, column=3, value=char.get("info", ""))
        _apply_cell_style(ws.cell(row=row, column=3), wrap=True)

        ws.cell(row=row, column=4, value=char.get("bio", ""))
        _apply_cell_style(ws.cell(row=row, column=4), wrap=True)

        ws.cell(row=row, column=5, value=char.get("prompt", ""))
        _apply_cell_style(ws.cell(row=row, column=5), wrap=True)

    # 列宽
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 60

    # 行高
    for row in range(2, len(characters) + 2):
        ws.row_dimensions[row].height = 200


# ─── 场景信息表 ─────────────────────────────────────────

def create_scene_sheet(ws, scenes: List[Dict[str, Any]]) -> None:
    """创建场景信息表工作表"""
    ws.title = "场景信息表"

    headers = ["场景编号", "场景名字", "场景提示词"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    for row, scene in enumerate(scenes, 2):
        ws.cell(row=row, column=1, value=scene.get("id", f"S{row-1:03d}"))
        _apply_cell_style(ws.cell(row=row, column=1), center=True)

        ws.cell(row=row, column=2, value=scene.get("name", ""))
        _apply_cell_style(ws.cell(row=row, column=2), center=True)

        ws.cell(row=row, column=3, value=scene.get("prompt", ""))
        _apply_cell_style(ws.cell(row=row, column=3), wrap=True)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 80

    for row in range(2, len(scenes) + 2):
        ws.row_dimensions[row].height = 120


# ─── 分镜信息表 ─────────────────────────────────────────

def create_storyboard_sheet(ws, storyboards: List[Dict[str, Any]]) -> None:
    """创建分镜信息表工作表（完整13列版本）"""
    ws.title = "分镜信息表"

    headers = [
        "第几集", "第几场", "第几个镜头", "时长(s)", "景别", "摄法",
        "画面内容", "台词/音效", "入镜角色", "场景标识",
        "首帧提示词", "尾帧提示词", "视频提示词",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)

    for row, shot in enumerate(storyboards, 2):
        ws.cell(row=row, column=1, value=shot.get("episode", 1))
        _apply_cell_style(ws.cell(row=row, column=1), center=True)

        ws.cell(row=row, column=2, value=shot.get("scene", 1))
        _apply_cell_style(ws.cell(row=row, column=2), center=True)

        ws.cell(row=row, column=3, value=shot.get("shot", 1))
        _apply_cell_style(ws.cell(row=row, column=3), center=True)

        ws.cell(row=row, column=4, value=shot.get("duration", ""))
        _apply_cell_style(ws.cell(row=row, column=4), center=True)

        ws.cell(row=row, column=5, value=shot.get("shot_size", ""))
        _apply_cell_style(ws.cell(row=row, column=5), center=True)

        ws.cell(row=row, column=6, value=shot.get("camera", ""))
        _apply_cell_style(ws.cell(row=row, column=6), center=True)

        ws.cell(row=row, column=7, value=shot.get("content", ""))
        _apply_cell_style(ws.cell(row=row, column=7), wrap=True)

        ws.cell(row=row, column=8, value=shot.get("dialogue", ""))
        _apply_cell_style(ws.cell(row=row, column=8), wrap=True)

        ws.cell(row=row, column=9, value=shot.get("characters", ""))
        _apply_cell_style(ws.cell(row=row, column=9), center=True)

        ws.cell(row=row, column=10, value=shot.get("scene_label", ""))
        _apply_cell_style(ws.cell(row=row, column=10), center=True)

        ws.cell(row=row, column=11, value=shot.get("first_frame", ""))
        _apply_cell_style(ws.cell(row=row, column=11), wrap=True)

        ws.cell(row=row, column=12, value=shot.get("last_frame", ""))
        _apply_cell_style(ws.cell(row=row, column=12), wrap=True)

        ws.cell(row=row, column=13, value=shot.get("video_prompt", ""))
        _apply_cell_style(ws.cell(row=row, column=13), wrap=True)

    # 列宽
    col_widths = [8, 8, 10, 8, 10, 10, 55, 30, 15, 15, 45, 45, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    for row in range(2, len(storyboards) + 2):
        ws.row_dimensions[row].height = 150


# ─── 主入口 ─────────────────────────────────────────────

def generate_excel_from_skill_data(
    characters: List[Dict[str, Any]],
    scenes: List[Dict[str, Any]],
    storyboards: List[Dict[str, Any]],
    output_dir: str = ".",
    project_name: str = "",
) -> Dict[str, Any]:
    """
    从技能结构化数据生成 Excel 文件。

    生成一个包含三个工作表的 Excel 文件：
    - 人物信息表
    - 场景信息表
    - 分镜信息表

    Args:
        characters: 人物数据列表
        scenes: 场景数据列表
        storyboards: 分镜数据列表
        output_dir: 输出目录
        project_name: 项目名称（用作文件名前缀）

    Returns:
        包含文件路径和统计信息的字典
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl 未安装，请运行 pip install openpyxl")

    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)

    # 文件名
    safe_name = project_name.replace(" ", "_").replace("/", "_") if project_name else f"漫剧_{int(time.time())}"
    filename = f"{safe_name}_完整表格.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb = Workbook()

    # 如果三个数据源都有，创建在一个文件中
    has_data = False

    # 人物信息表
    if characters:
        ws_char = wb.active if not has_data else wb.create_sheet()
        create_character_sheet(ws_char, characters)
        has_data = True
    else:
        # 确保默认 sheet 不残留
        if wb.active and wb.active.title == "Sheet":
            wb.active.title = "人物信息表"
            ws_char = wb.active
            # 即使没有数据也创建表头
            headers = ["人物编号", "人物姓名", "人物信息提取", "人物小传", "人物三视图提示词"]
            for col, header in enumerate(headers, 1):
                cell = ws_char.cell(row=1, column=col, value=header)
                _apply_header_style(cell)
            ws_char.column_dimensions["A"].width = 10
            ws_char.column_dimensions["B"].width = 15
            ws_char.column_dimensions["C"].width = 45
            ws_char.column_dimensions["D"].width = 50
            ws_char.column_dimensions["E"].width = 60
        has_data = True

    # 场景信息表
    ws_scene = wb.create_sheet()
    create_scene_sheet(ws_scene, scenes if scenes else [])

    # 分镜信息表
    ws_story = wb.create_sheet()
    create_storyboard_sheet(ws_story, storyboards if storyboards else [])

    wb.save(output_path)

    episodes = set(s.get("episode", 1) for s in storyboards) if storyboards else set()

    result = {
        "file_path": output_path,
        "file_name": filename,
        "characters_count": len(characters),
        "scenes_count": len(scenes),
        "storyboards_count": len(storyboards),
        "episodes_count": len(episodes),
    }

    logger.info(
        "[ExcelGenerator] Excel 生成完成: %s (角色=%d, 场景=%d, 分镜=%d, 集=%d)",
        output_path, len(characters), len(scenes), len(storyboards), len(episodes),
    )

    return result
